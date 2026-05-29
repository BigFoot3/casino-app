"""
Import joueurs depuis users_pass.xlsx.
Usage : cd /root/casino && source venv/bin/activate && python scripts/import_players.py
"""

import sys
import os
import secrets
import bcrypt
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

XLSX_PATH      = os.path.join(ROOT, 'users_pass.xlsx')
DEFAULT_TOKENS = 100

# Colonnes (1-based, openpyxl)
COL_PRENOM   = 1
COL_NOM      = 2
COL_JETONS   = 4
COL_USERNAME = 5
COL_MDP      = 6


def main():
    # Initialise l'app (init_db + migrations)
    from app import create_app
    create_app()

    from db import db_conn

    print(f"Fichier : {XLSX_PATH}", flush=True)
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    assert ws is not None, "Impossible d'accéder à la feuille active du classeur"
    print(f"Feuille : {ws.title} — {ws.max_row - 1} ligne(s) à traiter\n", flush=True)

    created = 0
    skipped = 0
    errors  = 0

    for row_idx in range(2, ws.max_row + 1):
        prenom_val = ws.cell(row_idx, COL_PRENOM).value
        nom_val    = ws.cell(row_idx, COL_NOM).value

        if not prenom_val or not nom_val:
            continue

        prenom = str(prenom_val).strip()
        nom    = str(nom_val).strip()
        if not prenom or not nom:
            continue

        username = f"{prenom.capitalize()}{nom.capitalize()}"

        jetons_val = ws.cell(row_idx, COL_JETONS).value
        try:
            jetons = int(str(jetons_val)) if jetons_val is not None else DEFAULT_TOKENS
        except (ValueError, TypeError):
            jetons = DEFAULT_TOKENS

        mdp     = secrets.token_urlsafe(10)
        pw_hash = bcrypt.hashpw(mdp.encode(), bcrypt.gensalt(rounds=10)).decode()

        try:
            with db_conn() as conn:
                conn.execute("BEGIN IMMEDIATE")
                result = conn.execute(
                    "INSERT OR IGNORE INTO users (username, password_hash, role, tokens)"
                    " VALUES (?, ?, 'player', ?)",
                    (username, pw_hash, jetons),
                )
                if result.rowcount == 0:
                    conn.execute("ROLLBACK")
                    print(f"⚠️  SKIP: {username} existe déjà", flush=True)
                    skipped += 1
                else:
                    conn.execute("COMMIT")
                    ws.cell(row_idx, COL_USERNAME).value = username
                    ws.cell(row_idx, COL_MDP).value      = mdp
                    print(f"✅ {username} créé ({jetons} jetons)", flush=True)
                    created += 1

        except Exception as exc:
            print(f"❌ Erreur pour {username} : {exc}", flush=True)
            errors += 1

    wb.save(XLSX_PATH)
    print(f"\n✅ {created} créés | ⚠️  {skipped} ignorés (déjà existants) | ❌ {errors} erreurs",
          flush=True)


if __name__ == "__main__":
    main()
